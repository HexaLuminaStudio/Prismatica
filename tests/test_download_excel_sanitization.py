from openpyxl import load_workbook

from app.core.services.global_download import GlobalDownloadWorker
from app.core.services.hsk_download import HSKDownloadWorker
from app.core.utils.excel import sanitizeExcelCellValue


ILLEGAL_CONTROL_CHARACTERS = "".join(
    chr(codePoint)
    for codePoint in (*range(0x00, 0x09), 0x0B, 0x0C, *range(0x0E, 0x20))
)


def test_sanitize_excel_cell_value_removes_only_forbidden_controls():
    sourceText = f"甲\t乙\n丙\r丁{ILLEGAL_CONTROL_CHARACTERS}戊"

    assert sanitizeExcelCellValue(sourceText) == "甲\t乙\n丙\r丁戊"
    assert sanitizeExcelCellValue(123) == 123
    assert sanitizeExcelCellValue(None) is None


def test_hsk_download_saves_rows_containing_illegal_controls(tmp_path):
    outputPath = tmp_path / "hsk.xlsx"
    sourceSentence = (
        "粮食{CC食粮[F糧]}分布的不公平，是{CC1就是}我们每个人的“ 知足”观念"
        "{CD来}可以{CC2可}{CJ+sy得到}解决的，我想。"
    )
    worker = HSKDownloadWorker({"taskId": "hsk-excel-test"})
    data = [
        {
            "sentence": f"{sourceSentence[:18]}\x0b{sourceSentence[18:]}",
            "comment": "保留\t制表符、\n换行和\r回车",
        }
    ]

    assert worker._processDataToExcel(data, str(outputPath)) is True

    workbook = load_workbook(outputPath, read_only=True)
    worksheet = workbook.active
    assert worksheet["A2"].value == sourceSentence
    assert worksheet["B2"].value == "保留\t制表符、\n换行和\r回车"
    workbook.close()


def test_global_download_uses_same_prewrite_sanitization(tmp_path):
    outputPath = tmp_path / "global.xlsx"
    worker = GlobalDownloadWorker({"taskId": "global-excel-test"})

    assert worker._processDataToExcel(
        [{"text": f"前文{ILLEGAL_CONTROL_CHARACTERS}后文"}],
        str(outputPath),
    ) is True

    workbook = load_workbook(outputPath, read_only=True)
    worksheet = workbook.active
    assert worksheet["A2"].value == "前文后文"
    workbook.close()
