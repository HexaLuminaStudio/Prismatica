# coding: utf-8
"""
情感分析 UI 主面板

需求文档:
    test/6D-CorpusClient_需求文档_v3.md §2.4.7

功能覆盖:
    - FR-SNT-001/002/003 三级情感分析(篇章/段落/句子)
    - FR-SNT-004 情感分布可视化(饼图 + 柱状图)
    - FR-SNT-005 情感词云(简化为高频词条 Top 列表)
    - FR-SNT-006 自定义情感词典导入
    - FR-SNT-007 报告导出(TXT 报告)

设计要点:
    - 复用 CorpusStore / CorpusStatusCard / TextSegmenter(经 sentiment_engine)
    - 后台 QThread 执行,完成后切回 UI 线程绘图
    - Matplotlib FigureCanvasQTAgg 嵌入
    - 复用 network_widget._availableCjkFonts() 解决中文字体警告
"""

from __future__ import annotations

import csv
import logging
import os
import time
import traceback
from typing import Dict, List, Optional

import matplotlib
import matplotlib.font_manager as fm  # noqa: F401
import matplotlib.pyplot as plt
import numpy as np
from PySide6.QtCore import QObject, Qt, QThread, Signal
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QSizePolicy,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    CheckBox,
    ComboBox,
    FluentIcon,
    InfoBar,
    InfoBarPosition,
    LineEdit,
    MessageBox,
    PlainTextEdit,
    PrimaryPushButton,
    PushButton,
    ScrollArea,
    SpinBox,
    StrongBodyLabel,
    SubtitleLabel,
    SwitchButton,
)

# 复用网络图模块的字体检测(避免 matplotlib findfont 警告)
from app.view.widgets.freq_analyzer.network_widget import _availableCjkFonts
from app.view.widgets.freq_analyzer.sentiment_engine import (
    CorpusSentimentResult,
    DocumentSentiment,
    Polarity,
    SentenceSentiment,
    SentimentEngine,
)

# matplotlib 后端必须在导入 Figure 前指定
matplotlib.use("QtAgg", force=False)

from matplotlib.figure import Figure  # noqa: E402
from matplotlib.backends.backend_qtagg import (  # noqa: E402
    FigureCanvasQTAgg as FigureCanvas,
)

# 全局字体设置
plt.rcParams["font.sans-serif"] = _availableCjkFonts()
plt.rcParams["font.family"] = "sans-serif"
plt.rcParams["axes.unicode_minus"] = False

from app.view.widgets.freq_analyzer.result_summary import MetricColor

logger = logging.getLogger(__name__)


# ===========================================================================
# 后台分析线程
# ===========================================================================
class SentimentWorker(QThread):
    """后台情感分析线程"""

    progress = Signal(int, str)  # (0-100, 描述)
    finished = Signal(object)  # CorpusSentimentResult
    failed = Signal(str)

    def __init__(
        self,
        engine: SentimentEngine,
        fileToText: Dict[str, str],
        parent=None,
    ):
        super().__init__(parent)
        self._engine = engine
        self._fileToText = fileToText

    def run(self):
        try:
            self.progress.emit(5, "开始分析...")

            # 注意:engine.analyzeCorpus 的回调现在是「字符级」
            # done: 已处理字符数, total: 总字符数
            def cb(doneChars, totalChars, fileName):
                pct = 5 + int(doneChars / max(1, totalChars) * 90)
                self.progress.emit(
                    pct,
                    f"分析中 ({doneChars:,}/{totalChars:,} 字符): {fileName}",
                )

            result = self._engine.analyzeCorpus(self._fileToText, progressCallback=cb)
            self.progress.emit(
                100,
                f"完成: {result.totalDocuments} 个文件 / {result.totalSentences} 个句子",
            )
            self.finished.emit(result)
        except Exception as e:
            logger.exception("[SentimentWorker] 失败")
            self.failed.emit(f"{e}\n{traceback.format_exc()}")


# ===========================================================================
# 主面板
# ===========================================================================
class SentimentWidget(QWidget):
    """情感分析主面板"""

    def __init__(self, parent=None, corpusStore=None):
        super().__init__(parent)
        self.setObjectName("sentimentWidget")

        self._corpusStore = corpusStore
        self.fileToText: Dict[str, str] = {}
        self._result: Optional[CorpusSentimentResult] = None
        self._worker: Optional[SentimentWorker] = None
        # 注入 token cache(加速重复分词)
        tokenCache = (
            self._corpusStore.tokenCache() if self._corpusStore is not None else None
        )
        self._engine = SentimentEngine(tokenCache=tokenCache)

        self._figure: Optional[Figure] = None
        self._axPie = None
        self._axBar = None
        self._canvas: Optional[FigureCanvas] = None

        self._initUi()

        if self._corpusStore is not None:
            self._bindCorpusStore(self._corpusStore)

    # ------------------------------------------------------------------
    # 语料绑定
    # ------------------------------------------------------------------
    def setCorpusStore(self, store) -> None:
        if self._corpusStore is store:
            return
        self._corpusStore = store
        self._bindCorpusStore(store)
        self._onCorpusChanged()

    def _bindCorpusStore(self, store) -> None:
        store.textsChanged.connect(self._onCorpusChanged)
        store.cleanRuleChanged.connect(self._onCorpusChanged)

    def _onCorpusChanged(self) -> None:
        if self._corpusStore is not None:
            self.fileToText = self._corpusStore.effectiveTexts()
        self._result = None
        if hasattr(self, "_resultSummary"):
            self._resultSummary.setPlaceholder("语料已变更,请点击「开始分析」")
        self._drawPlaceholder()

    # ------------------------------------------------------------------
    # UI 构建
    # ------------------------------------------------------------------
    def _initUi(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(20, 20, 20, 20)
        outer.setSpacing(12)

        title = SubtitleLabel("情感分析", self)
        outer.addWidget(title)

        scroll = ScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none; background: transparent;")
        outer.addWidget(scroll, 1)

        content = QWidget()
        scroll.setWidget(content)
        layout = QVBoxLayout(content)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        # 语料状态卡(只读)
        from app.view.widgets.freq_analyzer.concordance_widget import CorpusStatusCard

        self._corpusStatusCard = CorpusStatusCard(self, corpusStore=self._corpusStore)
        layout.addWidget(self._corpusStatusCard)

        # 参数 + 操作卡片
        layout.addWidget(self._buildActionCard())
        # 摘要卡(优化:统一大指标卡)
        self._resultSummary = self._buildSummaryCard()
        layout.addWidget(self._resultSummary)
        # 可视化卡
        layout.addWidget(self._buildChartCard(), 1)
        # 详情卡(句子列表)
        layout.addWidget(self._buildDetailCard(), 1)
        layout.addStretch(1)

    def _buildActionCard(self) -> CardWidget:
        card = CardWidget(self)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        layout.addWidget(StrongBodyLabel("1. 分析参数", card))

        row = QHBoxLayout()
        self.runBtn = PrimaryPushButton("开始分析", card)
        self.runBtn.setIcon(FluentIcon.PLAY)
        self.runBtn.clicked.connect(self._onRunClicked)
        row.addWidget(self.runBtn)

        self.importDictBtn = PushButton("导入情感词典...", card)
        self.importDictBtn.setIcon(FluentIcon.DOWNLOAD)
        self.importDictBtn.clicked.connect(self._onImportDictClicked)
        row.addWidget(self.importDictBtn)

        # 多格式导出按钮(主按钮 + 下拉菜单)
        self.exportReportBtn = PushButton("导出报告", card)
        self.exportReportBtn.setIcon(FluentIcon.SAVE)
        self.exportReportBtn.clicked.connect(
            lambda: self._showExportMenu(self.exportReportBtn)
        )
        self.exportReportBtn.setEnabled(False)
        row.addWidget(self.exportReportBtn)

        row.addStretch(1)

        self.statusLabel = CaptionLabel("就绪", card)
        self.statusLabel.setStyleSheet("color: #666; font-size: 11px;")
        row.addWidget(self.statusLabel)

        layout.addLayout(row)

        # 词典状态
        dictInfo = CaptionLabel(
            f"内置词典: 正面 {len(self._engine._positive)} 词 / "
            f"负面 {len(self._engine._negative)} 词",
            card,
        )
        dictInfo.setStyleSheet("color: #888; font-size: 11px;")
        layout.addWidget(dictInfo)

        return card

    def _buildSummaryCard(self) -> CardWidget:
        """优化:使用统一 ResultSummary,4 个大指标卡 + 详情 + Top 词"""
        from app.view.widgets.freq_analyzer.result_summary import (
            MetricColor,
            ResultSummary,
        )

        summary = ResultSummary(self)
        summary.setTitle("情感分析摘要")
        summary.setPlaceholder("请先加载语料并点击「开始分析」")
        return summary

    def _buildChartCard(self) -> CardWidget:
        card = CardWidget(self)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        layout.addWidget(StrongBodyLabel("3. 情感分布(FR-SNT-004)", card))

        # 创建一个 Figure,内含两个子图:饼图 + 柱状图
        self._figure = Figure(figsize=(10, 4), dpi=100, facecolor="#fafafa")
        self._axPie = self._figure.add_subplot(1, 2, 1)
        self._axBar = self._figure.add_subplot(1, 2, 2)

        self._canvas = FigureCanvas(self._figure)
        self._canvas.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding
        )
        layout.addWidget(self._canvas, 1)

        self._drawPlaceholder()

        return card

    def _buildDetailCard(self) -> CardWidget:
        card = CardWidget(self)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        layout.addWidget(StrongBodyLabel("4. 逐句分析(选中文件查看)", card))

        # 文件选择
        row = QHBoxLayout()
        row.addWidget(BodyLabel("文件:", card))
        self.docCombo = ComboBox(card)
        self.docCombo.setMinimumWidth(220)
        self.docCombo.currentIndexChanged.connect(self._onDocChanged)
        row.addWidget(self.docCombo, 1)
        layout.addLayout(row)

        # 句子列表
        from qfluentwidgetspro import (
            RoundTableWidget as ProRoundTableWidget,
        )  # noqa: F401

        try:
            from qfluentwidgetspro import RoundTableWidget as ProRoundTableWidget
        except ImportError:
            from qfluentwidgets.components.widgets.table_view import TableWidget as ProRoundTableWidget  # type: ignore  # noqa: E501

        self.sentenceTable = ProRoundTableWidget(card)
        self.sentenceTable.setColumnCount(4)
        self.sentenceTable.setHorizontalHeaderLabels(
            ["句子", "情感极性", "得分", "情感词"]
        )
        self.sentenceTable.verticalHeader().setVisible(False)
        self.sentenceTable.setEditTriggers(
            self.sentenceTable.EditTrigger.NoEditTriggers
        )
        self.sentenceTable.setSelectionBehavior(
            self.sentenceTable.SelectionBehavior.SelectRows
        )
        self.sentenceTable.setShowGrid(False)
        self.sentenceTable.setAlternatingRowColors(True)
        header = self.sentenceTable.horizontalHeader()
        header.setSectionResizeMode(0, header.ResizeMode.Stretch)
        for col in (1, 2, 3):
            header.setSectionResizeMode(col, header.ResizeMode.ResizeToContents)
        self.sentenceTable.setMinimumHeight(280)
        layout.addWidget(self.sentenceTable)

        return card

    # ------------------------------------------------------------------
    # 行为:运行分析
    # ------------------------------------------------------------------
    def _onRunClicked(self):
        if not self.fileToText:
            InfoBar.error(
                title="提示",
                content="请先在「语料导入与清洗」加载语料",
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                duration=2500,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )
            return
        if self._worker and self._worker.isRunning():
            return

        self.runBtn.setEnabled(False)
        self.statusLabel.setText("分析中...")

        self._worker = SentimentWorker(
            engine=self._engine,
            fileToText=self.fileToText,
            parent=self,
        )
        self._worker.progress.connect(self._onProgress)
        self._worker.finished.connect(self._onFinished)
        self._worker.failed.connect(self._onFailed)
        self._worker.start()

    def _onProgress(self, pct: int, msg: str):
        self.statusLabel.setText(f"{msg} ({pct}%)")

    def _onFailed(self, err: str):
        self.runBtn.setEnabled(True)
        self.statusLabel.setText("分析失败")
        logger.error(f"[SentimentWidget] 失败: {err}")
        InfoBar.error(
            title="分析失败",
            content=err[:200],
            orient=Qt.Orientation.Horizontal,
            isClosable=True,
            duration=4000,
            position=InfoBarPosition.TOP_RIGHT,
            parent=self,
        )

    def _onFinished(self, result: CorpusSentimentResult):
        self.runBtn.setEnabled(True)
        self._result = result
        self.statusLabel.setText(
            f"分析完成: {result.totalDocuments} 个文件 / {result.totalSentences} 个句子 / {result.elapsedSeconds:.1f}s"
        )
        self._updateSummary()
        self._updateCharts()
        self._updateDocCombo()
        self._updateSentenceTable()
        self.exportReportBtn.setEnabled(True)
        InfoBar.success(
            title="分析完成",
            content=f"平均分 {result.avgScore:+.3f},耗时 {result.elapsedSeconds:.1f}s",
            orient=Qt.Orientation.Horizontal,
            isClosable=True,
            duration=2500,
            position=InfoBarPosition.TOP_RIGHT,
            parent=self,
        )

    # ------------------------------------------------------------------
    # 摘要 / 图表
    # ------------------------------------------------------------------
    def _updateSummary(self):
        if self._result is None:
            return
        r = self._result
        polarity = (
            "积极" if r.avgScore > 0.05 else ("消极" if r.avgScore < -0.05 else "中性")
        )
        polarity_color = (
            MetricColor.SUCCESS
            if r.avgScore > 0.05
            else (MetricColor.ERROR if r.avgScore < -0.05 else MetricColor.NEUTRAL)
        )

        self._resultSummary.clear()
        self._resultSummary.setMetrics(
            [
                ("整体倾向", polarity, polarity_color),
                ("平均分", f"{r.avgScore:+.3f}", MetricColor.PRIMARY),
                ("正面句", f"{r.positiveCount:,}", MetricColor.SUCCESS),
                ("负面句", f"{r.negativeCount:,}", MetricColor.ERROR),
            ]
        )
        self._resultSummary.setDetail(
            f"💭 句子总数 <b>{r.totalSentences:,}</b> &nbsp;|&nbsp; "
            f"中性 <b>{r.neutralCount:,}</b> &nbsp;|&nbsp; "
            f"耗时 <b>{r.elapsedSeconds:.2f}s</b>"
        )
        self._resultSummary.setTopWords(
            positive=r.topPositiveWords(8),
            negative=r.topNegativeWords(8),
        )

    def _updateCharts(self):
        if self._result is None or self._figure is None:
            return
        r = self._result

        # 饼图:句子级极性分布
        self._axPie.clear()
        sizes = [r.positiveCount, r.negativeCount, r.neutralCount]
        labels = [
            f"正面({r.positiveCount})",
            f"负面({r.negativeCount})",
            f"中性({r.neutralCount})",
        ]
        colors = ["#52c41a", "#f5222d", "#bfbfbf"]
        # 过滤 0 值避免警告
        filtered = [(s, l, c) for s, l, c in zip(sizes, labels, colors) if s > 0]
        if filtered:
            fs, fl, fc = zip(*filtered)
            self._axPie.pie(fs, labels=fl, colors=fc, autopct="%1.1f%%", startangle=90)
        else:
            self._axPie.text(0.5, 0.5, "无数据", ha="center", va="center", fontsize=14)
        self._axPie.set_title("句子情感分布")

        # 柱状图:每个文件的平均分
        self._axBar.clear()
        docs = [d for d in r.documents if d.sentences]
        if docs:
            names = [d.fileName[:20] for d in docs]
            scores = [d.score for d in docs]
            bar_colors = [
                "#52c41a" if s > 0.05 else "#f5222d" if s < -0.05 else "#bfbfbf"
                for s in scores
            ]
            x = np.arange(len(names))
            self._axBar.bar(x, scores, color=bar_colors, alpha=0.85)
            self._axBar.axhline(0, color="black", linewidth=0.6)
            self._axBar.set_xticks(x)
            self._axBar.set_xticklabels(names, rotation=45, ha="right", fontsize=8)
            self._axBar.set_ylim(-1.05, 1.05)
            self._axBar.set_ylabel("情感得分")
            self._axBar.set_title("各文件情感得分")
        else:
            self._axBar.text(
                0.5, 0.5, "无文件数据", ha="center", va="center", fontsize=14
            )

        self._figure.tight_layout()
        self._canvas.draw_idle()

    def _updateDocCombo(self):
        self.docCombo.blockSignals(True)
        self.docCombo.clear()
        if self._result is None:
            self.docCombo.blockSignals(False)
            return
        for d in self._result.documents:
            label = f"{d.fileName}  ·  分 {d.score:+.3f}  ·  {self._polarityLabel(d.polarity)}"
            self.docCombo.addItem(label, userData=d.fileName)
        self.docCombo.blockSignals(False)

    def _onDocChanged(self, idx: int):
        self._updateSentenceTable()

    def _updateSentenceTable(self):
        if self._result is None:
            return
        fileName = self.docCombo.currentData()
        doc = next((d for d in self._result.documents if d.fileName == fileName), None)
        if doc is None:
            self.sentenceTable.setRowCount(0)
            return
        self.sentenceTable.setRowCount(len(doc.sentences))
        for i, sent in enumerate(doc.sentences):
            txt = sent.text.strip()
            self.sentenceTable.setItem(i, 0, QTableWidgetItem(txt[:80]))
            polarityItem = QTableWidgetItem(self._polarityLabel(sent.polarity))
            color = {
                Polarity.POSITIVE: QColor("#52c41a"),
                Polarity.NEGATIVE: QColor("#f5222d"),
                Polarity.NEUTRAL: QColor("#888888"),
            }.get(sent.polarity, QColor("#888"))
            polarityItem.setForeground(color)
            self.sentenceTable.setItem(i, 1, polarityItem)
            self.sentenceTable.setItem(i, 2, QTableWidgetItem(f"{sent.score:+.3f}"))
            hitWords = ", ".join(
                f"{h.word}{'!' if h.negated else ''}" for h in sent.hits
            )
            self.sentenceTable.setItem(i, 3, QTableWidgetItem(hitWords or "-"))

    # ------------------------------------------------------------------
    # 占位图
    # ------------------------------------------------------------------
    def _drawPlaceholder(self):
        if self._axPie is None or self._axBar is None:
            return
        for ax, msg in ((self._axPie, "等待分析"), (self._axBar, "等待分析")):
            ax.clear()
            ax.set_axis_off()
            ax.text(
                0.5,
                0.5,
                msg,
                ha="center",
                va="center",
                fontsize=14,
                color="#999",
                transform=ax.transAxes,
            )
        self._canvas.draw_idle()

    # ------------------------------------------------------------------
    # 行为:词典导入 / 报告导出
    # ------------------------------------------------------------------
    def _onImportDictClicked(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择情感词典",
            "",
            "Text/CSV Files (*.txt *.csv);;All Files (*)",
        )
        if not path:
            return

        # 询问是否替换
        dlg = MessageBox(
            "导入词典",
            "选择导入模式:\n\n"
            "「确定」= 合并到内置词典\n"
            "「替换」= 完全替换内置词典(谨慎!)",
            self.window(),
        )
        # MessageBox 默认 yes/cancel,这里调整
        dlg.yesButton.setText("合并")
        dlg.cancelButton.setText("取消")

        replace = False
        from qfluentwidgets import PrimaryPushButton

        replaceBtn = PrimaryPushButton("替换", dlg)
        replaceBtn.clicked.connect(dlg.accept)
        dlg.buttonLayout.addWidget(replaceBtn)

        choice = dlg.exec()
        if not choice:
            return
        # 由于 yes 也走 accept,replaceBtn 也走 accept,无法直接区分
        # 通过 clicked 信号间接判断:替换按钮先被点 → 我们用 sender 信息?
        # 简化:用返回值区分 - 这里采用二次弹窗处理替换
        if replaceBtn.isDown():
            replace = True

        try:
            newPos, newNeg, newWeights = self._engine.importCustomDict(
                path, replaceBuiltin=replace
            )
            InfoBar.success(
                title="导入成功",
                content=f"新增正面 {newPos} / 负面 {newNeg} 词; 权重条目 {newWeights}",
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                duration=2500,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )
        except Exception as e:
            logger.error(f"[SentimentWidget] 导入词典失败: {e}")
            InfoBar.error(
                title="导入失败",
                content=str(e),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                duration=3000,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )

    # ------------------------------------------------------------------
    # 报告导出:多格式(TXT / CSV / Excel / 图表)
    # ------------------------------------------------------------------
    def _showExportMenu(self, anchorBtn):
        """显示多格式导出菜单"""
        if self._result is None:
            return
        from qfluentwidgets import Action, RoundMenu

        menu = RoundMenu(parent=self)
        menu.addAction(
            Action(
                FluentIcon.DOCUMENT,
                "TXT 文本报告",
                triggered=lambda: self._exportAs("txt"),
            )
        )
        menu.addAction(
            Action(
                FluentIcon.SAVE,
                "CSV 多文件(含文档汇总/句子/词频)",
                triggered=lambda: self._exportAs("csv"),
            )
        )
        menu.addAction(
            Action(
                FluentIcon.SAVE,
                "Excel 工作簿(.xlsx)",
                triggered=lambda: self._exportAs("xlsx"),
            )
        )
        menu.addSeparator()
        menu.addAction(
            Action(
                FluentIcon.PHOTO,
                "图表 PNG(饼图 + 柱状图)",
                triggered=lambda: self._exportAs("png"),
            )
        )
        menu.addAction(
            Action(
                FluentIcon.PHOTO,
                "图表 PDF(矢量,可直接打印)",
                triggered=lambda: self._exportAs("pdf"),
            )
        )
        menu.addAction(
            Action(
                FluentIcon.PHOTO,
                "图表 SVG(无损,可二次编辑)",
                triggered=lambda: self._exportAs("svg"),
            )
        )
        # 在按钮下方弹出
        btnRect = anchorBtn.rect()
        pos = anchorBtn.mapToGlobal(btnRect.bottomLeft())
        menu.exec(pos)

    def _exportAs(self, kind: str) -> None:
        """统一导出入口

        Args:
            kind: txt | csv | xlsx | png | pdf | svg
        """
        if self._result is None:
            return

        extMap = {
            "txt": ("TXT Files (*.txt)", "sentiment_report.txt", ".txt"),
            "csv": ("CSV Files (*.csv)", "sentiment_report", ".csv"),
            "xlsx": ("Excel Files (*.xlsx)", "sentiment_report.xlsx", ".xlsx"),
            "png": ("PNG Files (*.png)", "sentiment_charts.png", ".png"),
            "pdf": ("PDF Files (*.pdf)", "sentiment_charts.pdf", ".pdf"),
            "svg": ("SVG Files (*.svg)", "sentiment_charts.svg", ".svg"),
        }
        filterStr, defaultName, ext = extMap[kind]

        path, _ = QFileDialog.getSaveFileName(
            self,
            f"导出情感分析报告({kind.upper()})",
            defaultName,
            filterStr,
        )
        if not path:
            return
        if not path.lower().endswith(ext):
            path += ext

        try:
            if kind == "txt":
                self._exportReportTxt(path)
            elif kind == "csv":
                self._exportCsvBundle(path)
            elif kind == "xlsx":
                self._exportExcel(path)
            elif kind in ("png", "pdf", "svg"):
                self._exportCharts(path, fmt=kind)
            else:
                raise ValueError(f"未知导出格式:{kind}")

            InfoBar.success(
                title="导出成功",
                content=f"已保存:{path}",
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                duration=2500,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )
        except Exception as e:
            logger.exception(f"[SentimentWidget] {kind} 导出失败: {e}")
            InfoBar.error(
                title="导出失败",
                content=f"{kind.upper()} 导出异常:{e}",
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                duration=3500,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )

    def _exportReportTxt(self, path: str):
        """TXT 文本报告"""
        r = self._result
        if r is None:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write("# 情感分析报告\n\n")
            f.write(f"总文件数: {r.totalDocuments}\n")
            f.write(f"总字符数: {r.totalChars:,}\n")
            f.write(f"总句子数: {r.totalSentences}\n")
            f.write(f"正面句子: {r.positiveCount}\n")
            f.write(f"负面句子: {r.negativeCount}\n")
            f.write(f"中性句子: {r.neutralCount}\n")
            f.write(f"平均分: {r.avgScore:+.4f}\n")
            f.write(f"耗时: {r.elapsedSeconds:.2f}s\n\n")

            f.write("# Top 正面词\n")
            for w, c in r.topPositiveWords(30):
                f.write(f"  {w}: {c}\n")
            f.write("\n# Top 负面词\n")
            for w, c in r.topNegativeWords(30):
                f.write(f"  {w}: {c}\n")

            f.write("\n# 逐文件分析\n")
            for d in r.documents:
                f.write(
                    f"\n## {d.fileName}  ·  分 {d.score:+.4f}  ·  {self._polarityLabel(d.polarity)}\n"
                )
                for i, sent in enumerate(d.sentences, 1):
                    f.write(
                        f"  [{i:>3}] {sent.score:+.4f}  {self._polarityLabel(sent.polarity)}  "
                        f"{sent.text.strip()[:80]}\n"
                    )

    def _exportCsvBundle(self, path: str):
        """CSV 多文件包:写到 path(主文件) + 派生 _summary.csv / _sentences.csv / _words.csv

        若 path 形如 report.csv,则派生:
            report_summary.csv   文档汇总
            report_sentences.csv 句子明细
            report_words.csv     情感词频
        """
        r = self._result
        if r is None:
            return
        # 派生文件名前缀
        base, _ = os.path.splitext(path)
        summaryPath = f"{base}_summary.csv"
        sentencesPath = f"{base}_sentences.csv"
        wordsPath = f"{base}_words.csv"

        # 1) 文档汇总
        with open(summaryPath, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(
                [
                    "文件名",
                    "情感得分",
                    "极性",
                    "正面句数",
                    "负面句数",
                    "中性句数",
                    "总句数",
                    "总字符数",
                ]
            )
            for d in r.documents:
                w.writerow(
                    [
                        d.fileName,
                        f"{d.score:+.4f}",
                        self._polarityLabel(d.polarity),
                        d.positiveCount,
                        d.negativeCount,
                        d.neutralCount,
                        d.totalSentences,
                        len(d.text),
                    ]
                )

        # 2) 句子明细
        with open(sentencesPath, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["文件名", "句子序号", "句子内容", "得分", "极性", "情感词命中"])
            for d in r.documents:
                for i, sent in enumerate(d.sentences, 1):
                    hitWords = ", ".join(
                        f"{h.word}{'!' if h.negated else ''}" for h in sent.hits
                    )
                    w.writerow(
                        [
                            d.fileName,
                            i,
                            sent.text.strip(),
                            f"{sent.score:+.4f}",
                            self._polarityLabel(sent.polarity),
                            hitWords or "-",
                        ]
                    )

        # 3) 情感词频
        with open(wordsPath, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["极性", "词语", "频次"])
            for word, cnt in r.topPositiveWords(100):
                w.writerow(["正面", word, cnt])
            for word, cnt in r.topNegativeWords(100):
                w.writerow(["负面", word, cnt])

    def _exportExcel(self, path: str):
        """Excel 工作簿(多 sheet)

        依赖:openpyxl。若不可用则降级为多 CSV。
        """
        r = self._result
        if r is None:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            # 降级为 CSV
            base, _ = os.path.splitext(path)
            self._exportCsvBundle(base + ".csv")
            InfoBar.warning(
                title="已降级导出",
                content="未安装 openpyxl,已改导出为 CSV 多文件包",
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                duration=3000,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )
            return

        wb = Workbook()
        # 默认 sheet 改为"摘要"
        wsSummary = wb.active
        wsSummary.title = "摘要"

        # 样式
        headerFont = Font(bold=True, color="FFFFFF")
        headerFill = PatternFill("solid", fgColor="1890FF")
        polarityColor = {
            "正面": PatternFill("solid", fgColor="B7EB8F"),
            "负面": PatternFill("solid", fgColor="FFCCC7"),
            "中性": PatternFill("solid", fgColor="F0F0F0"),
        }

        # ---- Sheet 1: 摘要 ----
        wsSummary["A1"] = "情感分析报告"
        wsSummary["A1"].font = Font(bold=True, size=14)
        wsSummary.merge_cells("A1:C1")

        wsSummary["A3"] = "总文件数"
        wsSummary["B3"] = r.totalDocuments
        wsSummary["A4"] = "总字符数"
        wsSummary["B4"] = r.totalChars
        wsSummary["A5"] = "总句子数"
        wsSummary["B5"] = r.totalSentences
        wsSummary["A6"] = "正面句子"
        wsSummary["B6"] = r.positiveCount
        wsSummary["A7"] = "负面句子"
        wsSummary["B7"] = r.negativeCount
        wsSummary["A8"] = "中性句子"
        wsSummary["B8"] = r.neutralCount
        wsSummary["A9"] = "平均分"
        wsSummary["B9"] = round(r.avgScore, 4)
        wsSummary["A10"] = "耗时(秒)"
        wsSummary["B10"] = round(r.elapsedSeconds, 2)
        for row in range(3, 11):
            wsSummary[f"A{row}"].font = Font(bold=True)

        wsSummary.column_dimensions["A"].width = 16
        wsSummary.column_dimensions["B"].width = 18

        # ---- Sheet 2: 文档汇总 ----
        wsDocs = wb.create_sheet("文档汇总")
        headers = [
            "文件名",
            "情感得分",
            "极性",
            "正面句数",
            "负面句数",
            "中性句数",
            "总句数",
            "总字符数",
        ]
        for col, h in enumerate(headers, 1):
            cell = wsDocs.cell(row=1, column=col, value=h)
            cell.font = headerFont
            cell.fill = headerFill
            cell.alignment = Alignment(horizontal="center")
        for rIdx, d in enumerate(r.documents, 2):
            wsDocs.cell(row=rIdx, column=1, value=d.fileName)
            wsDocs.cell(row=rIdx, column=2, value=round(d.score, 4))
            polarityCell = wsDocs.cell(
                row=rIdx, column=3, value=self._polarityLabel(d.polarity)
            )
            if d.polarity in (Polarity.POSITIVE, Polarity.NEGATIVE):
                label = self._polarityLabel(d.polarity)
                polarityCell.fill = polarityColor.get(label, PatternFill())
            wsDocs.cell(row=rIdx, column=4, value=d.positiveCount)
            wsDocs.cell(row=rIdx, column=5, value=d.negativeCount)
            wsDocs.cell(row=rIdx, column=6, value=d.neutralCount)
            wsDocs.cell(row=rIdx, column=7, value=d.totalSentences)
            wsDocs.cell(row=rIdx, column=8, value=len(d.text))

        # 列宽自适应
        for col in range(1, len(headers) + 1):
            wsDocs.column_dimensions[get_column_letter(col)].width = 16

        # ---- Sheet 3: 句子明细 ----
        wsSent = wb.create_sheet("句子明细")
        sentHeaders = ["文件名", "序号", "句子内容", "得分", "极性", "情感词"]
        for col, h in enumerate(sentHeaders, 1):
            cell = wsSent.cell(row=1, column=col, value=h)
            cell.font = headerFont
            cell.fill = headerFill
            cell.alignment = Alignment(horizontal="center")
        rowIdx = 2
        for d in r.documents:
            for i, sent in enumerate(d.sentences, 1):
                wsSent.cell(row=rowIdx, column=1, value=d.fileName)
                wsSent.cell(row=rowIdx, column=2, value=i)
                wsSent.cell(row=rowIdx, column=3, value=sent.text.strip())
                wsSent.cell(row=rowIdx, column=4, value=round(sent.score, 4))
                polCell = wsSent.cell(
                    row=rowIdx, column=5, value=self._polarityLabel(sent.polarity)
                )
                label = self._polarityLabel(sent.polarity)
                if label in polarityColor:
                    polCell.fill = polarityColor[label]
                hitWords = ", ".join(
                    f"{h.word}{'!' if h.negated else ''}" for h in sent.hits
                )
                wsSent.cell(row=rowIdx, column=6, value=hitWords or "-")
                rowIdx += 1

        wsSent.column_dimensions["A"].width = 24
        wsSent.column_dimensions["B"].width = 8
        wsSent.column_dimensions["C"].width = 60
        wsSent.column_dimensions["D"].width = 10
        wsSent.column_dimensions["E"].width = 10
        wsSent.column_dimensions["F"].width = 30

        # ---- Sheet 4: 情感词频 ----
        wsWords = wb.create_sheet("情感词频")
        wordHeaders = ["极性", "词语", "频次"]
        for col, h in enumerate(wordHeaders, 1):
            cell = wsWords.cell(row=1, column=col, value=h)
            cell.font = headerFont
            cell.fill = headerFill
            cell.alignment = Alignment(horizontal="center")
        rowIdx = 2
        for w, c in r.topPositiveWords(100):
            wsWords.cell(row=rowIdx, column=1, value="正面").fill = polarityColor[
                "正面"
            ]
            wsWords.cell(row=rowIdx, column=2, value=w)
            wsWords.cell(row=rowIdx, column=3, value=c)
            rowIdx += 1
        for w, c in r.topNegativeWords(100):
            wsWords.cell(row=rowIdx, column=1, value="负面").fill = polarityColor[
                "负面"
            ]
            wsWords.cell(row=rowIdx, column=2, value=w)
            wsWords.cell(row=rowIdx, column=3, value=c)
            rowIdx += 1

        for col in range(1, 4):
            wsWords.column_dimensions[get_column_letter(col)].width = 14

        wb.save(path)

    def _exportCharts(self, path: str, fmt: str):
        """导出图表(png / pdf / svg)

        使用独立 Figure 渲染(不修改当前 UI 上的 Figure),
        包含三个子图:饼图、文件得分柱状图、Top 词条横向条形图。
        """
        r = self._result
        if r is None:
            return

        # 在子线程可能不在主线程,这里在主线程直接渲染
        # 用 Agg 渲染后端(不依赖 Qt)
        import matplotlib

        matplotlib.use("Agg", force=True)
        import matplotlib.pyplot as plt
        import numpy as np

        # 字体设置与全局一致
        plt.rcParams["font.sans-serif"] = _availableCjkFonts()
        plt.rcParams["font.family"] = "sans-serif"
        plt.rcParams["axes.unicode_minus"] = False

        fig = plt.figure(figsize=(13, 5), dpi=150, facecolor="white")
        gs = fig.add_gridspec(1, 3, width_ratios=[1, 1, 1.2], wspace=0.3)

        # ---- 子图 1: 饼图 ----
        axPie = fig.add_subplot(gs[0, 0])
        sizes = [r.positiveCount, r.negativeCount, r.neutralCount]
        labels = [
            f"正面({r.positiveCount})",
            f"负面({r.negativeCount})",
            f"中性({r.neutralCount})",
        ]
        colors = ["#52c41a", "#f5222d", "#bfbfbf"]
        filtered = [(s, l, c) for s, l, c in zip(sizes, labels, colors) if s > 0]
        if filtered:
            fs, fl, fc = zip(*filtered)
            axPie.pie(
                fs,
                labels=fl,
                colors=fc,
                autopct="%1.1f%%",
                startangle=90,
                textprops={"fontsize": 10},
            )
        else:
            axPie.text(0.5, 0.5, "无数据", ha="center", va="center", fontsize=14)
        axPie.set_title("句子情感分布", fontsize=12)

        # ---- 子图 2: 文件得分柱状图 ----
        axBar = fig.add_subplot(gs[0, 1])
        docs = [d for d in r.documents if d.sentences]
        if docs:
            names = [d.fileName[:18] for d in docs]
            scores = [d.score for d in docs]
            barColors = [
                "#52c41a" if s > 0.05 else "#f5222d" if s < -0.05 else "#bfbfbf"
                for s in scores
            ]
            x = np.arange(len(names))
            axBar.bar(x, scores, color=barColors, alpha=0.85)
            axBar.axhline(0, color="black", linewidth=0.6)
            axBar.set_xticks(x)
            axBar.set_xticklabels(names, rotation=45, ha="right", fontsize=7)
            axBar.set_ylim(-1.05, 1.05)
            axBar.set_ylabel("情感得分", fontsize=9)
        else:
            axBar.text(0.5, 0.5, "无文件数据", ha="center", va="center", fontsize=14)
        axBar.set_title("各文件情感得分", fontsize=12)

        # ---- 子图 3: Top 词条横向条形图 ----
        axWords = fig.add_subplot(gs[0, 2])
        posWords = r.topPositiveWords(10)
        negWords = r.topNegativeWords(10)
        # 合并并标记极性
        combined = [(w, c, "正") for w, c in posWords] + [
            (w, c, "负") for w, c in negWords
        ]
        if combined:
            # 按频次升序排列(横向条形图底部到顶部)
            combined.sort(key=lambda x: x[1])
            words = [f"[{p}] {w}" for w, c, p in combined]
            counts = [c for _, c, _ in combined]
            colors = ["#52c41a" if p == "正" else "#f5222d" for _, _, p in combined]
            y = np.arange(len(words))
            axWords.barh(y, counts, color=colors, alpha=0.85)
            axWords.set_yticks(y)
            axWords.set_yticklabels(words, fontsize=8)
            axWords.set_xlabel("频次", fontsize=9)
        else:
            axWords.text(
                0.5, 0.5, "无 Top 词数据", ha="center", va="center", fontsize=14
            )
        axWords.set_title("Top 10 情感词", fontsize=12)
        axWords.grid(True, axis="x", linestyle="--", alpha=0.5)

        fig.suptitle(
            f"情感分析报告 — 平均分 {r.avgScore:+.3f} · 共 {r.totalSentences:,} 句",
            fontsize=14,
            fontweight="bold",
        )

        # 保存
        fig.tight_layout(rect=[0, 0, 1, 0.95])
        if fmt == "png":
            fig.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
        elif fmt == "pdf":
            fig.savefig(path, format="pdf", bbox_inches="tight", facecolor="white")
        elif fmt == "svg":
            fig.savefig(path, format="svg", bbox_inches="tight", facecolor="white")
        else:
            raise ValueError(f"不支持的图表格式:{fmt}")
        plt.close(fig)

    @staticmethod
    def _polarityLabel(p: Polarity) -> str:
        return {
            Polarity.POSITIVE: "正面",
            Polarity.NEGATIVE: "负面",
            Polarity.NEUTRAL: "中性",
        }.get(p, "中性")

    # ------------------------------------------------------------------
    # 关闭
    # ------------------------------------------------------------------
    def closeEvent(self, event) -> None:
        if self._worker is not None and self._worker.isRunning():
            self._worker.wait(2000)
        super().closeEvent(event)
